import io
import os
from typing import List
from openpyxl import Workbook
from openpyxl.styles import Font as ExcelFont, Alignment, PatternFill, Border, Side
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfbase.pdfmetrics import registerFontFamily

from app.models.report import ItemCategory
from app.schemas.report import ActivityReportResponse

# --- Turkish UTF-8 Font Registration for ReportLab PDF ---
FONT_NAME = "Helvetica"
FONT_BOLD_NAME = "Helvetica-Bold"

try:
    win_font = "C:\\Windows\\Fonts\\arial.ttf"
    win_bold = "C:\\Windows\\Fonts\\arialbd.ttf"
    win_italic = "C:\\Windows\\Fonts\\ariali.ttf"
    win_bolditalic = "C:\\Windows\\Fonts\\arialbi.ttf"

    if os.path.exists(win_font):
        pdfmetrics.registerFont(TTFont("Arial", win_font))
        FONT_NAME = "Arial"
        
        if os.path.exists(win_bold):
            pdfmetrics.registerFont(TTFont("Arial-Bold", win_bold))
            FONT_BOLD_NAME = "Arial-Bold"
        else:
            pdfmetrics.registerFont(TTFont("Arial-Bold", win_font))
            
        if os.path.exists(win_italic):
            pdfmetrics.registerFont(TTFont("Arial-Italic", win_italic))
        else:
            pdfmetrics.registerFont(TTFont("Arial-Italic", win_font))
            
        if os.path.exists(win_bolditalic):
            pdfmetrics.registerFont(TTFont("Arial-BoldItalic", win_bolditalic))
        else:
            pdfmetrics.registerFont(TTFont("Arial-BoldItalic", win_font))
            
        registerFontFamily(
            'Arial',
            normal='Arial',
            bold='Arial-Bold',
            italic='Arial-Italic',
            boldItalic='Arial-BoldItalic'
        )
except Exception:
    FONT_NAME = "Helvetica"
    FONT_BOLD_NAME = "Helvetica-Bold"


class ExportService:
    """
    Faaliyet raporlarını Excel (.xlsx) ve PDF (.pdf) formatlarına
    dönüştürerek dinamik dosya akışı (BytesIO) oluşturan servis sınıfı.
    """

    @staticmethod
    def get_unit_hierarchy_parts(unit_id: int, unit_map: dict) -> tuple[str, str]:
        department_name = "-"
        directorate_name = "-"
        
        if not unit_id or unit_id not in unit_map:
            return department_name, directorate_name
            
        current = unit_map[unit_id]
        subunit_name = None
        for _ in range(10):
            if not current:
                break
                
            u_type = getattr(current, "unit_type", None)
            if hasattr(u_type, "value"):
                u_type = u_type.value
                
            if u_type == "DEPARTMENT" and department_name == "-":
                department_name = current.name
            elif u_type == "DIRECTORATE" and directorate_name == "-":
                directorate_name = current.name
            elif u_type == "SUB_UNIT" and not subunit_name and current.id == unit_id:
                subunit_name = current.name
                
            parent_id = getattr(current, "parent_id", None)
            current = unit_map.get(parent_id) if parent_id else None
            
        if subunit_name:
            if directorate_name == "-":
                directorate_name = subunit_name
            else:
                directorate_name = f"{directorate_name} / {subunit_name}"
                
        if department_name == "-" and directorate_name == "-":
            directorate_name = unit_map[unit_id].name
            
        return department_name, directorate_name

    @staticmethod
    async def export_reports_to_excel(reports: List[ActivityReportResponse], db) -> io.BytesIO:
        """
        Faaliyet raporları listesini kategori bazlı 3 ayrı sayfada (tab)
        ve tarih sırasına göre (en eskiden yeniye) Excel dosyasına dönüştürür.
        """
        from app.models.unit import Unit
        from sqlalchemy import select
        
        # Tüm birimleri çek ve map'le
        result = await db.execute(select(Unit))
        all_units = result.scalars().all()
        unit_map = {u.id: u for u in all_units}

        # Girdileri tarih olarak en eskiden yeniye doğru sırala (Yıl, Ay, Rapor ID)
        reports.sort(key=lambda r: (r.year, r.month, r.id))

        wb = Workbook()
        
        # Sayfaları Oluştur
        ws_yapilan = wb.active
        ws_yapilan.title = "Yapılan İşler"
        
        ws_yapilacak = wb.create_sheet(title="Yapılacak İşler")
        ws_kordinasyon = wb.create_sheet(title="Koordinasyon Gerektiren İşler")

        # Stil Tanımlamaları
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = ExcelFont(name="Calibri", size=11, bold=True, color="FFFFFF")
        data_font = ExcelFont(name="Calibri", size=10)
        center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        left_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

        thin_border = Border(
            left=Side(style="thin", color="D3D3D3"),
            right=Side(style="thin", color="D3D3D3"),
            top=Side(style="thin", color="D3D3D3"),
            bottom=Side(style="thin", color="D3D3D3")
        )

        headers_std = [
            "Rapor ID", "Yıl / Ay", "Personel", "Ünvan", "Daire Başkanlığı", "Şube Md. / Alt Birim", 
            "Açıklama / İçerik"
        ]
        headers_kord = [
            "Rapor ID", "Yıl / Ay", "Personel", "Ünvan", "Daire Başkanlığı", "Şube Md. / Alt Birim", 
            "Koordinasyon Gerektiren İş", "İlgili Kurum Kuruluşlar", "Çözüm Önerileri"
        ]

        # Başlıkları Yaz ve Stilleri Uygula
        for ws, headers in [(ws_yapilan, headers_std), (ws_yapilacak, headers_std), (ws_kordinasyon, headers_kord)]:
            ws.append(headers)
            for col_num in range(1, len(headers) + 1):
                cell = ws.cell(row=1, column=col_num)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_alignment

        # Veri Yazma Takipçileri (Row index)
        rows = {
            ItemCategory.YAPILAN_ISLER: 2,
            ItemCategory.YAPILACAK_ISLER: 2,
            ItemCategory.KORDINASYON_GEREKTIREN_ISLER: 2
        }

        sheets = {
            ItemCategory.YAPILAN_ISLER: ws_yapilan,
            ItemCategory.YAPILACAK_ISLER: ws_yapilacak,
            ItemCategory.KORDINASYON_GEREKTIREN_ISLER: ws_kordinasyon
        }

        for report in reports:
            period = f"{report.year} / {report.month:02d}"

            all_items = getattr(report, "yapilan_isler", []) + getattr(report, "yapilacak_isler", []) + getattr(report, "koordinasyon_isleri", [])
            if all_items:
                for item in all_items:
                    target_ws = sheets.get(item.category)
                    if not target_ws:
                        continue
                        
                    user_name = item.creator.full_name if item.creator and item.creator.full_name else "N/A"
                    title = item.creator.title if item.creator and item.creator.title else "-"
                    department, directorate = ExportService.get_unit_hierarchy_parts(item.creator.unit_id if item.creator else None, unit_map)
                    
                    curr_row = rows[item.category]
                    
                    if item.category == ItemCategory.KORDINASYON_GEREKTIREN_ISLER:
                        insts = item.related_institutions or "-"
                        sols = item.solution_proposals or "-"
                        row_data = [
                            report.id, period, user_name, title, department, directorate,
                            item.content, insts, sols
                        ]
                        target_ws.append(row_data)
                        for col_idx in range(1, len(row_data) + 1):
                            cell = target_ws.cell(row=curr_row, column=col_idx)
                            cell.font = data_font
                            cell.border = thin_border
                            cell.alignment = center_alignment if col_idx in [1, 2] else left_alignment
                    else:
                        row_data = [
                            report.id, period, user_name, title, department, directorate,
                            item.content
                        ]
                        target_ws.append(row_data)
                        for col_idx in range(1, len(row_data) + 1):
                            cell = target_ws.cell(row=curr_row, column=col_idx)
                            cell.font = data_font
                            cell.border = thin_border
                            cell.alignment = center_alignment if col_idx in [1, 2] else left_alignment
                            
                    rows[item.category] += 1

        # Sütun Genişliklerini Ayarla
        column_widths_std = {
            "A": 12, "B": 12, "C": 22, "D": 20, 
            "E": 28, "F": 28, "G": 50
        }
        column_widths_kord = {
            "A": 12, "B": 12, "C": 22, "D": 20, 
            "E": 28, "F": 28, "G": 40, "H": 30, "I": 30
        }

        for col_letter, width in column_widths_std.items():
            ws_yapilan.column_dimensions[col_letter].width = width
            ws_yapilacak.column_dimensions[col_letter].width = width

        for col_letter, width in column_widths_kord.items():
            ws_kordinasyon.column_dimensions[col_letter].width = width

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    @staticmethod
    def generate_excel_template() -> io.BytesIO:
        wb = Workbook()
        
        ws_yapilan = wb.active
        ws_yapilan.title = "Yapılan İşler"
        ws_yapilacak = wb.create_sheet(title="Yapılacak İşler")
        ws_kordinasyon = wb.create_sheet(title="Koordinasyon Gerektiren İşler")

        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = ExcelFont(name="Calibri", size=11, bold=True, color="FFFFFF")
        center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        headers_std = [
            "Rapor ID", "Yıl / Ay", "Personel", "Ünvan", "Daire Başkanlığı", "Şube Md. / Alt Birim", 
            "Açıklama / İçerik"
        ]
        headers_kord = [
            "Rapor ID", "Yıl / Ay", "Personel", "Ünvan", "Daire Başkanlığı", "Şube Md. / Alt Birim", 
            "Koordinasyon Gerektiren İş", "İlgili Kurum Kuruluşlar", "Çözüm Önerileri"
        ]

        for ws, headers in [(ws_yapilan, headers_std), (ws_yapilacak, headers_std), (ws_kordinasyon, headers_kord)]:
            ws.append(headers)
            for col_num in range(1, len(headers) + 1):
                cell = ws.cell(row=1, column=col_num)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_alignment
                
        # widths
        for ws in [ws_yapilan, ws_yapilacak]:
            ws.column_dimensions["A"].width = 10
            ws.column_dimensions["B"].width = 15
            ws.column_dimensions["C"].width = 25
            ws.column_dimensions["D"].width = 25
            ws.column_dimensions["E"].width = 30
            ws.column_dimensions["F"].width = 30
            ws.column_dimensions["G"].width = 80
            
        ws_kordinasyon.column_dimensions["A"].width = 10
        ws_kordinasyon.column_dimensions["B"].width = 15
        ws_kordinasyon.column_dimensions["C"].width = 25
        ws_kordinasyon.column_dimensions["D"].width = 25
        ws_kordinasyon.column_dimensions["E"].width = 30
        ws_kordinasyon.column_dimensions["F"].width = 30
        ws_kordinasyon.column_dimensions["G"].width = 40
        ws_kordinasyon.column_dimensions["H"].width = 30
        ws_kordinasyon.column_dimensions["I"].width = 40

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    @staticmethod
    async def export_report_to_pdf(report: ActivityReportResponse, db) -> io.BytesIO:
        """
        Tekil bir faaliyet raporunu PDF belgesine dönüştürür.
        """
        from app.models.unit import Unit
        from sqlalchemy import select
        
        # Tüm birimleri çek ve map'le
        result = await db.execute(select(Unit))
        all_units = result.scalars().all()
        unit_map = {u.id: u for u in all_units}

        output = io.BytesIO()
        doc = SimpleDocTemplate(
            output,
            pagesize=A4,
            rightMargin=30,
            leftMargin=30,
            topMargin=30,
            bottomMargin=30
        )

        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            "ReportTitle",
            parent=styles["Heading1"],
            fontName=FONT_BOLD_NAME,
            fontSize=16,
            leading=20,
            alignment=1,
            textColor=colors.HexColor("#1F4E78"),
            spaceAfter=15
        )
        subtitle_style = ParagraphStyle(
            "ReportSubTitle",
            parent=styles["Normal"],
            fontName=FONT_NAME,
            fontSize=10,
            leading=14,
            textColor=colors.HexColor("#333333"),
            spaceAfter=10
        )
        cat_header_style = ParagraphStyle(
            "CatHeader",
            parent=styles["Heading2"],
            fontName=FONT_BOLD_NAME,
            fontSize=12,
            leading=16,
            textColor=colors.HexColor("#1F4E78"),
            spaceBefore=10,
            spaceAfter=5
        )
        body_style = ParagraphStyle(
            "BodyTextCustom",
            parent=styles["Normal"],
            fontName=FONT_NAME,
            fontSize=9,
            leading=12,
            textColor=colors.HexColor("#222222")
        )

        elements = []

        # Başlık
        period_str = f"{report.year} Yılı {report.month}. Ay Faaliyet Raporu"
        elements.append(Paragraph(f"<b>FAALİYET RAPORU</b>", title_style))
        elements.append(Paragraph(f"<b>Dönem:</b> {period_str}", subtitle_style))

        # Kullanıcı ve Birim Bilgileri Tablosu
        department, directorate = ExportService.get_unit_hierarchy_parts(report.user.unit.id if report.user and report.user.unit else None, unit_map)

        info_data = [
            [Paragraph("<b>Daire Başkanlığı:</b>", body_style), Paragraph(department, body_style),
             Paragraph("<b>Şube Md. / Birim:</b>", body_style), Paragraph(directorate, body_style)]
        ]

        info_table = Table(info_data, colWidths=[100, 200, 100, 130])
        info_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor("#F2F4F8")),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('INNERGRID', (0, 0), (-1, -1), 0.5, colors.HexColor("#D3D3D3")),
            ('BOX', (0, 0), (-1, -1), 1, colors.HexColor("#1F4E78")),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ]))
        elements.append(info_table)
        elements.append(Spacer(1, 15))

        # Kategoriler ve Faaliyetler
        categories = [
            (ItemCategory.YAPILAN_ISLER, "1. YAPILAN İŞLER"),
            (ItemCategory.YAPILACAK_ISLER, "2. YAPILACAK İŞLER"),
            (ItemCategory.KORDINASYON_GEREKTIREN_ISLER, "3. KOORDİNASYON GEREKTİREN İŞLER")
        ]

        for cat_enum, cat_title in categories:
            elements.append(Paragraph(cat_title, cat_header_style))
            all_items = getattr(report, "yapilan_isler", []) + getattr(report, "yapilacak_isler", []) + getattr(report, "koordinasyon_isleri", [])
            cat_items = [item for item in all_items if item.category == cat_enum]

            if not cat_items:
                elements.append(Paragraph("<i>Bu kategoride kayıtlı faaliyet bulunmamaktadır.</i>", body_style))
                elements.append(Spacer(1, 8))
                continue

            if cat_enum == ItemCategory.KORDINASYON_GEREKTIREN_ISLER:
                table_data = [[
                    Paragraph("<b>#</b>", body_style), 
                    Paragraph("<b>Personel</b>", body_style),
                    Paragraph("<b>Ünvan</b>", body_style),
                    Paragraph("<b>Koordinasyon Gerektiren İş</b>", body_style),
                    Paragraph("<b>İlgili/İlişkili Kurum Kuruluşlar</b>", body_style),
                    Paragraph("<b>Çözüm Önerileri</b>", body_style)
                ]]
                for idx, item in enumerate(cat_items, 1):
                    insts = item.related_institutions or "-"
                    sols = item.solution_proposals or "-"
                    user_name = item.creator.full_name if item.creator and item.creator.full_name else "N/A"
                    title = item.creator.title if item.creator and item.creator.title else "-"
                    table_data.append([
                        Paragraph(str(idx), body_style),
                        Paragraph(user_name, body_style),
                        Paragraph(title, body_style),
                        Paragraph(item.content.replace("\n", "<br/>") if item.content else "-", body_style),
                        Paragraph(insts.replace("\n", "<br/>"), body_style),
                        Paragraph(sols.replace("\n", "<br/>"), body_style)
                    ])
                cat_table = Table(table_data, colWidths=[20, 70, 70, 130, 120, 120])
            else:
                table_data = [[
                    Paragraph("<b>#</b>", body_style), 
                    Paragraph("<b>Personel</b>", body_style),
                    Paragraph("<b>Ünvan</b>", body_style),
                    Paragraph("<b>Faaliyet / İş Açıklaması</b>", body_style)
                ]]
                for idx, item in enumerate(cat_items, 1):
                    user_name = item.creator.full_name if item.creator and item.creator.full_name else "N/A"
                    title = item.creator.title if item.creator and item.creator.title else "-"
                    table_data.append([
                        Paragraph(str(idx), body_style),
                        Paragraph(user_name, body_style),
                        Paragraph(title, body_style),
                        Paragraph(item.content.replace("\n", "<br/>") if item.content else "-", body_style)
                    ])
                cat_table = Table(table_data, colWidths=[20, 80, 80, 350])

            cat_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#E6EEF8")),
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ('INNERGRID', (0, 0), (-1, -1), 0.5, colors.HexColor("#E0E0E0")),
                ('BOX', (0, 0), (-1, -1), 0.5, colors.HexColor("#1F4E78")),
                ('TOPPADDING', (0, 0), (-1, -1), 5),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
            ]))
            elements.append(cat_table)
            elements.append(Spacer(1, 10))

        doc.build(elements)
        output.seek(0)
        return output