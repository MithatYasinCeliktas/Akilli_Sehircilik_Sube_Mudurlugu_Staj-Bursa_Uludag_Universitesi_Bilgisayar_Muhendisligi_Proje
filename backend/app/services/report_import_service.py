import io
from typing import List, Dict, Any, Optional
from datetime import datetime
from openpyxl import load_workbook
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, and_, or_
from sqlalchemy.orm import selectinload

from app.models.user import User
from app.models.institution import Institution
import difflib
from app.models.unit import Unit
from app.models.report import ActivityReport, ReportItem, ItemCategory, ReportStatus
from app.core.exceptions import BadRequestException, ForbiddenException, NotFoundException
from app.services.report_service import ReportService

class ReportImportService:
    def __init__(self, db: AsyncSession, report_service: ReportService):
        self.db = db
        self.report_service = report_service

    async def _resolve_user(self, personnel: str, title: str, dept: str, div: str, current: User, subordinate_users: List[User]) -> Dict[str, Any]:
        """
        Attempts to resolve the target user based on provided fields using an in-memory list.
        Returns a dictionary with 'target_user_id', 'conflicts' (list of dicts), or 'error'.
        """
        # If all empty, it's the current user
        if not personnel and not title and not dept and not div:
            return {"target_user_id": current.id, "conflicts": []}

        if not subordinate_users:
            from app.models.user import UserRole
            if current.role == UserRole.USER:
                return {"error": "Sadece kendi adınıza rapor girebilirsiniz."}
            return {"error": "Alt çalışanınız bulunmamaktadır. Başkası adına rapor giremezsiniz."}

        # Filter users in the hierarchy in memory
        users = []
        personnel_lower = personnel.lower() if personnel else ""
        title_lower = title.lower() if title else ""
        
        for u in subordinate_users:
            match = True
            if personnel_lower and personnel_lower not in u.full_name.lower():
                match = False
            if title_lower and u.title and title_lower not in u.title.lower():
                match = False
            if match:
                users.append(u)

        # Akıllı eşleştirme: Eğer isim ve unvan ile arandığında sadece 1 kişi çıkıyorsa,
        # birim isimlerindeki (Şb. Md.lüğü vs.) yazım hatalarını görmezden gel ve direkt kabul et.
        if len(users) == 1:
            return {"target_user_id": users[0].id, "conflicts": []}

        # Further filter by unit names if provided (conflict resolution)
        filtered_users = []
        for u in users:
            match = True
            if dept and dept.strip():
                if not u.unit or dept.lower() not in u.unit.name.lower():
                    match = False
            if div and div.strip():
                if not u.unit or div.lower() not in u.unit.name.lower():
                    match = False
            if match:
                filtered_users.append(u)

        if not filtered_users:
            # Maybe they typed their own info?
            if personnel and personnel.lower() in current.full_name.lower():
                return {"target_user_id": current.id, "conflicts": []}
            
            from app.models.user import UserRole
            if current.role == UserRole.USER:
                return {"error": "Sadece kendi adınıza rapor girebilirsiniz."}
                
            return {"error": "Girilen bilgilere sahip, size bağlı bir personel bulunamadı."}
        
        if len(filtered_users) == 1:
            return {"target_user_id": filtered_users[0].id, "conflicts": []}
        
        # Conflicts
        conflicts = [
            {
                "id": u.id,
                "full_name": u.full_name,
                "title": u.title,
                "unit_name": u.unit.name if u.unit else "-"
            }
            for u in filtered_users
        ]
        return {"target_user_id": None, "conflicts": conflicts}

    async def preview_import(self, file_bytes: bytes, current_user: User, target: str = "OWN_REPORT") -> Dict[str, Any]:
        try:
            wb = load_workbook(filename=io.BytesIO(file_bytes), data_only=True, read_only=True)
        except Exception:
            raise BadRequestException("Geçersiz Excel dosyası. Lütfen xlsx formatında yükleyin.")

        result = {
            "valid": True,
            "global_errors": [],
            "rows": []
        }

        # Expected headers mapped to logical names
        expected_headers = {
            "Rapor ID": "id",
            "Yıl / Ay": "year_month",
            "Personel": "personnel",
            "Ünvan": "title",
            "Daire Başkanlığı": "dept",
            "Şube Md. / Alt Birim": "div",
            "Açıklama / İçerik": "content",
            "Koordinasyon Gerektiren İş": "kordinasyon",
            "İlgili Kurum Kuruluşlar": "kurum",
            "Çözüm Önerileri": "cozum"
        }
        
        base_keys = ["Rapor ID", "Yıl / Ay", "Personel", "Ünvan", "Daire Başkanlığı", "Şube Md. / Alt Birim"]
        
        
        # --- CACHING INSTITUTIONS ---
        query_insts = select(Institution).where(Institution.is_active == True)
        result_insts = await self.db.execute(query_insts)
        active_institutions = result_insts.scalars().all()
        active_inst_names = [i.name for i in active_institutions]
        active_inst_names_lower = {i.name.lower(): i.name for i in active_institutions}

        # --- CACHING FOR OPTIMIZATION ---
        subordinate_ids = await self.report_service.get_subordinate_ids_recursive(current_user.id)
        subordinate_users = []
        if subordinate_ids:
            query_users = select(User).options(selectinload(User.unit)).where(User.id.in_(subordinate_ids))
            result_users = await self.db.execute(query_users)
            subordinate_users = result_users.scalars().all()

        target_report_owner_id = current_user.manager_id if target == "MANAGER_REPORT" else current_user.id
        
        cached_reports_by_ym = {}
        if target_report_owner_id:
            query_reports = select(ActivityReport).where(ActivityReport.user_id == target_report_owner_id)
            result_reports = await self.db.execute(query_reports)
            all_target_reports = result_reports.scalars().all()
            for r in all_target_reports:
                ym_key = f"{r.year}-{r.month}"
                if ym_key not in cached_reports_by_ym:
                    cached_reports_by_ym[ym_key] = []
                cached_reports_by_ym[ym_key].append(r)

        row_index = 0
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            category = None
            if "yapılan" in sheet_name.lower():
                category = ItemCategory.YAPILAN_ISLER.value
                sheet_expected_keys = base_keys + ["Açıklama / İçerik"]
            elif "yapılacak" in sheet_name.lower():
                category = ItemCategory.YAPILACAK_ISLER.value
                sheet_expected_keys = base_keys + ["Açıklama / İçerik"]
            elif "koordinasyon" in sheet_name.lower() or "kordinasyon" in sheet_name.lower():
                category = ItemCategory.KORDINASYON_GEREKTIREN_ISLER.value
                sheet_expected_keys = base_keys + ["Koordinasyon Gerektiren İş", "İlgili Kurum Kuruluşlar", "Çözüm Önerileri"]
            else:
                continue

            empty_row_count = 0
            headers = []
            for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
                if row_idx == 0:
                    headers = [str(cell).strip() if cell else "" for cell in row][:len(sheet_expected_keys)]
                    
                    if headers != sheet_expected_keys:
                        result["global_errors"].append(f"'{sheet_name}' sayfasındaki sütun başlıkları şablonla birebir eşleşmiyor. Beklenen: {', '.join(sheet_expected_keys)}")
                        result["valid"] = False
                        return result
                    continue

                row_index += 1
                row_data = dict(zip(headers, row))
                
                # Sadece whitespace veya None içeren satırları boş say
                if not any(v for v in row_data.values() if v is not None and str(v).strip() != ""):
                    empty_row_count += 1
                    if empty_row_count > 20:
                        break  # 20 boş satır gördüysek tablonun sonuna gelmişizdir (Excel 1 milyon satır bug'ını engeller)
                    continue
                else:
                    empty_row_count = 0

                parsed_row = {
                    "original_index": row_index,
                    "sheet_name": sheet_name,
                    "category": category,
                    "id": row_data.get("Rapor ID"),
                    "year_month": row_data.get("Yıl / Ay"),
                    "personnel": row_data.get("Personel"),
                    "title": row_data.get("Ünvan"),
                    "dept": row_data.get("Daire Başkanlığı"),
                    "div": row_data.get("Şube Md. / Alt Birim"),
                    "content": row_data.get("Açıklama / İçerik"),
                    "kordinasyon": row_data.get("Koordinasyon Gerektiren İş"),
                    "kurum": row_data.get("İlgili Kurum Kuruluşlar"),
                    "cozum": row_data.get("Çözüm Önerileri"),
                    "errors": [],
                    "warnings": [],
                    "target_user_id": None,
                    "conflicts": []
                }

                for k in ["personnel", "title", "dept", "div"]:
                    if parsed_row[k] is not None and str(parsed_row[k]).strip() in ["", "None"]:
                        parsed_row[k] = None

                ym = str(parsed_row["year_month"]).strip() if parsed_row["year_month"] else ""
                now = datetime.now()
                parsed_row["year"] = now.year
                parsed_row["month"] = now.month
                if ym and "/" in ym:
                    try:
                        y, m = ym.split("/")
                        parsed_row["year"] = int(y.strip())
                        parsed_row["month"] = int(m.strip())
                    except ValueError:
                        parsed_row["errors"].append("'Yıl / Ay' formatı geçersiz. Beklenen: YYYY/AA")

                res = await self._resolve_user(
                    parsed_row["personnel"], 
                    parsed_row["title"], 
                    parsed_row["dept"], 
                    parsed_row["div"], 
                    current_user,
                    subordinate_users
                )
                
                if "error" in res:
                    parsed_row["errors"].append(res["error"])
                else:
                    parsed_row["target_user_id"] = res["target_user_id"]
                    parsed_row["conflicts"] = res["conflicts"]
                
                if not target_report_owner_id:
                    parsed_row["errors"].append("Bağlı olduğunuz bir yönetici bulunmuyor.")
                else:
                    if parsed_row["id"]:
                        try:
                            rid = int(parsed_row["id"])
                            report = await self.report_service.get_report_by_id(rid, current_user)
                            if report.user_id != target_report_owner_id:
                                parsed_row["errors"].append(f"Rapor ID {rid}, eklenmesi gereken yetkili rapora ait değil.")
                            else:
                                parsed_row["available_reports"] = [{"id": report.id, "title": report.title or f"{report.year}/{report.month} Raporu"}]
                                parsed_row["selected_report_id"] = report.id
                        except Exception as e:
                            parsed_row["errors"].append(str(e))
                    else:
                        ym_key = f"{parsed_row['year']}-{parsed_row['month']}"
                        target_reports = cached_reports_by_ym.get(ym_key, [])
                        
                        if not target_reports:
                            parsed_row["warnings"].append(f"Hedef rapor bulunamadı ({parsed_row['year']}/{parsed_row['month']}). İçe aktarım sırasında otomatik oluşturulacak.")
                            parsed_row["selected_report_id"] = None
                        else:
                            parsed_row["available_reports"] = [{"id": r.id, "title": r.title or f"{r.year}/{r.month} Raporu"} for r in target_reports]
                            if len(target_reports) == 1:
                                parsed_row["selected_report_id"] = target_reports[0].id
                            else:
                                parsed_row["selected_report_id"] = None

                if category == ItemCategory.KORDINASYON_GEREKTIREN_ISLER.value:
                    if not parsed_row.get("kordinasyon") or not str(parsed_row["kordinasyon"]).strip():
                        parsed_row["errors"].append("'Koordinasyon Gerektiren İş' alanı boş bırakılamaz.")
                    if not parsed_row.get("kurum") or not str(parsed_row["kurum"]).strip():
                        parsed_row["errors"].append("'İlgili Kurum Kuruluşlar' alanı boş bırakılamaz.")
                    else:
                        kurum_raw = str(parsed_row["kurum"]).strip()
                        kurum_parts = [p.strip() for p in kurum_raw.split(',') if p.strip()]
                        parsed_row["institution_conflicts"] = parsed_row.get("institution_conflicts", [])
                        
                        for kp in kurum_parts:
                            kp_lower = kp.lower()
                            if kp_lower not in active_inst_names_lower:
                                suggestions = difflib.get_close_matches(kp, active_inst_names, n=3, cutoff=0.5)
                                parsed_row["institution_conflicts"].append({
                                    "raw": kp,
                                    "suggestions": suggestions
                                })
                                parsed_row["errors"].append(f"'{kp}' kurumu sistemde bulunamadı.")
                                
                    if not parsed_row.get("cozum") or not str(parsed_row["cozum"]).strip():
                        parsed_row["errors"].append("'Çözüm Önerileri' alanı boş bırakılamaz.")
                else:
                    if not parsed_row.get("content") or not str(parsed_row["content"]).strip():
                        parsed_row["errors"].append("'Açıklama / İçerik' alanı boş bırakılamaz.")

                if parsed_row["errors"] or parsed_row["conflicts"]:
                    result["valid"] = False
                
                result["rows"].append(parsed_row)

        wb.close()
        return result

    async def revalidate_import(self, data: Dict[str, Any], current_user: User) -> Dict[str, Any]:
        result = {
            "valid": True,
            "global_errors": data.get("global_errors", []),
            "rows": []
        }
        
        target = data.get("target", "OWN_REPORT")
        
        
        # --- CACHING INSTITUTIONS ---
        query_insts = select(Institution).where(Institution.is_active == True)
        result_insts = await self.db.execute(query_insts)
        active_institutions = result_insts.scalars().all()
        active_inst_names = [i.name for i in active_institutions]
        active_inst_names_lower = {i.name.lower(): i.name for i in active_institutions}

        # --- CACHING FOR OPTIMIZATION ---
        subordinate_ids = await self.report_service.get_subordinate_ids_recursive(current_user.id)
        subordinate_users = []
        if subordinate_ids:
            query_users = select(User).options(selectinload(User.unit)).where(User.id.in_(subordinate_ids))
            result_users = await self.db.execute(query_users)
            subordinate_users = result_users.scalars().all()

        target_report_owner_id = current_user.manager_id if target == "MANAGER_REPORT" else current_user.id
        
        cached_reports_by_ym = {}
        if target_report_owner_id:
            query_reports = select(ActivityReport).where(ActivityReport.user_id == target_report_owner_id)
            result_reports = await self.db.execute(query_reports)
            all_target_reports = result_reports.scalars().all()
            for r in all_target_reports:
                ym_key = f"{r.year}-{r.month}"
                if ym_key not in cached_reports_by_ym:
                    cached_reports_by_ym[ym_key] = []
                cached_reports_by_ym[ym_key].append(r)
        
        for parsed_row in data.get("rows", []):
            parsed_row["errors"] = []
            parsed_row["conflicts"] = []
            
            for k in ["personnel", "title", "dept", "div"]:
                if parsed_row.get(k) is not None and str(parsed_row[k]).strip() in ["", "None"]:
                    parsed_row[k] = None

            ym = str(parsed_row.get("year_month", "")).strip()
            now = datetime.now()
            parsed_row["year"] = now.year
            parsed_row["month"] = now.month
            if ym and "/" in ym:
                try:
                    y, m = ym.split("/")
                    parsed_row["year"] = int(y.strip())
                    parsed_row["month"] = int(m.strip())
                except ValueError:
                    parsed_row["errors"].append("'Yıl / Ay' formatı geçersiz. Beklenen: YYYY/AA")

            res = await self._resolve_user(
                parsed_row.get("personnel"), 
                parsed_row.get("title"), 
                parsed_row.get("dept"), 
                parsed_row.get("div"), 
                current_user,
                subordinate_users
            )
            
            if "error" in res:
                parsed_row["errors"].append(res["error"])
            else:
                parsed_row["target_user_id"] = res["target_user_id"]
                parsed_row["conflicts"] = res["conflicts"]
            
            if not target_report_owner_id:
                parsed_row["errors"].append("Bağlı olduğunuz bir yönetici bulunmuyor.")
            else:
                if parsed_row.get("id"):
                    try:
                        rid = int(parsed_row["id"])
                        report = await self.report_service.get_report_by_id(rid, current_user)
                        if report.user_id != target_report_owner_id:
                            parsed_row["errors"].append(f"Rapor ID {rid}, eklenmesi gereken yetkili rapora ait değil.")
                        else:
                            parsed_row["available_reports"] = [{"id": report.id, "title": report.title or f"{report.year}/{report.month} Raporu"}]
                            parsed_row["selected_report_id"] = report.id
                    except Exception as e:
                        parsed_row["errors"].append(str(e))
                else:
                    ym_key = f"{parsed_row['year']}-{parsed_row['month']}"
                    target_reports = cached_reports_by_ym.get(ym_key, [])
                    
                    if not target_reports:
                        parsed_row["warnings"].append(f"Hedef rapor bulunamadı ({parsed_row['year']}/{parsed_row['month']}). İçe aktarım sırasında otomatik oluşturulacak.")
                        parsed_row["selected_report_id"] = None
                    else:
                        parsed_row["available_reports"] = [{"id": r.id, "title": r.title or f"{r.year}/{r.month} Raporu"} for r in target_reports]
                        if len(target_reports) == 1:
                            parsed_row["selected_report_id"] = target_reports[0].id
                        else:
                            parsed_row["selected_report_id"] = None

            category = parsed_row.get("category")
            if category == ItemCategory.KORDINASYON_GEREKTIREN_ISLER.value:
                    if not parsed_row.get("kordinasyon") or not str(parsed_row["kordinasyon"]).strip():
                        parsed_row["errors"].append("'Koordinasyon Gerektiren İş' alanı boş bırakılamaz.")
                    if not parsed_row.get("kurum") or not str(parsed_row["kurum"]).strip():
                        parsed_row["errors"].append("'İlgili Kurum Kuruluşlar' alanı boş bırakılamaz.")
                    else:
                        kurum_raw = str(parsed_row["kurum"]).strip()
                        kurum_parts = [p.strip() for p in kurum_raw.split(',') if p.strip()]
                        parsed_row["institution_conflicts"] = parsed_row.get("institution_conflicts", [])
                        
                        for kp in kurum_parts:
                            kp_lower = kp.lower()
                            if kp_lower not in active_inst_names_lower:
                                suggestions = difflib.get_close_matches(kp, active_inst_names, n=3, cutoff=0.5)
                                parsed_row["institution_conflicts"].append({
                                    "raw": kp,
                                    "suggestions": suggestions
                                })
                                parsed_row["errors"].append(f"'{kp}' kurumu sistemde bulunamadı.")
                                
                    if not parsed_row.get("cozum") or not str(parsed_row["cozum"]).strip():
                        parsed_row["errors"].append("'Çözüm Önerileri' alanı boş bırakılamaz.")
            else:
                if not parsed_row.get("content") or not str(parsed_row["content"]).strip():
                    parsed_row["errors"].append("'Açıklama / İçerik' alanı boş bırakılamaz.")

            if parsed_row["errors"] or parsed_row["conflicts"]:
                result["valid"] = False
            
            result["rows"].append(parsed_row)

        return result

    async def execute_import(self, data: Dict[str, Any], current_user: User) -> Dict[str, Any]:
        rows = data.get("rows", [])
        success_count = 0
        error_count = 0
        errors = []

        from app.models.report import ReportItemProposal, ProposalStatus
        from app.models.notification import Notification

        subordinate_ids = await self.report_service.get_subordinate_ids_recursive(current_user.id)

        target = data.get("target", "OWN_REPORT")
        
        for idx, row in enumerate(rows):
            try:
                target_user_id = row.get("target_user_id")
                if not target_user_id:
                    raise BadRequestException("Hedef personel belirlenemedi (Çatışma çözülmemiş).")
                
                # Determine target report owner based on user selection
                if target == "MANAGER_REPORT":
                    target_report_owner_id = current_user.manager_id
                else:
                    target_report_owner_id = current_user.id

                if not target_report_owner_id:
                    raise BadRequestException("Bağlı olduğunuz bir yönetici bulunmuyor (Kendi adınıza ekleme yapamazsınız).")
                
                # Enforce rule: if employee, they can only upload for themselves
                if not subordinate_ids and target_user_id != current_user.id:
                    raise ForbiddenException("Sadece kendi adınıza faaliyet ekleyebilirsiniz.")

                selected_report_id = row.get("selected_report_id")
                report = None
                
                if selected_report_id:
                    report = await self.db.get(ActivityReport, int(selected_report_id))
                    if not report or report.user_id != target_report_owner_id:
                        raise NotFoundException("Rapor bulunamadı veya yetkiniz yok.")
                else:
                    # Fallback to year/month
                    year = row.get("year")
                    month = row.get("month")
                    if not year or not month:
                        raise BadRequestException("Ne rapor ID ne de yıl/ay bilgisi bulundu.")
                    
                    from sqlalchemy.future import select
                    query = select(ActivityReport).where(
                        ActivityReport.year == int(year),
                        ActivityReport.month == int(month),
                        ActivityReport.user_id == target_report_owner_id
                    )
                    res = await self.db.execute(query)
                    report = res.scalars().first()
                    
                    if not report:
                        # Rapor yoksa otomatik oluştur!
                        report = ActivityReport(
                            user_id=target_report_owner_id,
                            year=int(year),
                            month=int(month),
                            title=f"{year}/{month} Faaliyet Raporu",
                            status="PENDING",
                            yapilan_is_ids=[],
                            yapilacak_is_ids=[],
                            koordinasyon_is_ids=[]
                        )
                        self.db.add(report)
                        await self.db.flush()
                
                category = row.get("category")
                content = row.get("kordinasyon") if category == ItemCategory.KORDINASYON_GEREKTIREN_ISLER.value else row.get("content")
                
                if target_user_id == current_user.id:
                    # Kendi adına ekliyorsa doğrudan ReportItem
                    new_item = ReportItem(
                        report_id=report.id,
                        category=category,
                        content=content,
                        related_institutions=row.get("kurum"),
                        solution_proposals=row.get("cozum"),
                        creator_id=current_user.id,
                        status=ReportStatus.PENDING
                    )
                    self.db.add(new_item)
                    await self.db.flush()
                    
                    from sqlalchemy.orm.attributes import flag_modified
                    
                    if new_item.category == ItemCategory.YAPILAN_ISLER:
                        report.yapilan_is_ids = (report.yapilan_is_ids or []) + [new_item.id]
                        flag_modified(report, "yapilan_is_ids")
                    elif new_item.category == ItemCategory.YAPILACAK_ISLER:
                        report.yapilacak_is_ids = (report.yapilacak_is_ids or []) + [new_item.id]
                        flag_modified(report, "yapilacak_is_ids")
                    elif new_item.category == ItemCategory.KORDINASYON_GEREKTIREN_ISLER:
                        report.koordinasyon_is_ids = (report.koordinasyon_is_ids or []) + [new_item.id]
                        flag_modified(report, "koordinasyon_is_ids")
                else:
                    # Alt çalışan adına ekliyorsa ReportItemProposal oluştur
                    new_proposal = ReportItemProposal(
                        manager_report_id=report.id,
                        target_user_id=target_user_id,
                        creator_id=current_user.id,
                        category=category,
                        content=content,
                        related_institutions=row.get("kurum"),
                        solution_proposals=row.get("cozum"),
                        status=ProposalStatus.PENDING
                    )
                    self.db.add(new_proposal)
                    await self.db.flush()

                    # Bildirim oluştur
                    notif_msg = f"Yöneticiniz sizin adınıza bir faaliyet ekledi. Onayınız bekleniyor. (Faaliyet Raporu ID: {report.id})"
                    new_notif = Notification(
                        user_id=target_user_id,
                        message=notif_msg,
                        is_read=False,
                        type="PROPOSAL_PENDING",
                        reference_id=new_proposal.id
                    )
                    self.db.add(new_notif)
                
                await self.db.commit()
                success_count += 1
            except Exception as e:
                await self.db.rollback()
                error_count += 1
                errors.append(f"Satır {row.get('original_index')}: {str(e)}")
        
        return {
            "success": True,
            "success_count": success_count,
            "error_count": error_count,
            "errors": errors
        }
